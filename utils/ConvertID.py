# -*- coding: utf-8 -*-
"""
Created on Sun Apr 10 20:54:38 2016

@author: nebula
"""

import openpyxl


class ConvertID:
    def __init__(self, xlsfilename_vib='vibrism.xlsx', xlsfilename_tx='braintx.xlsx'):
        wb_vib = openpyxl.load_workbook(xlsfilename_vib)
        
        ws_vib = wb_vib.get_sheet_by_name(u'行')
        self.cx2entrez = {}
        self.entrez2cx = {}
        for i in range(2, ws_vib.max_row):
            entrez = ws_vib['H'+str(i)].value
            cxid = ws_vib['E'+str(i)].value
            intensity = 0
            if entrez is not None and entrez != u'':
                entrez = int(entrez)
                self.cx2entrez[cxid] = int(entrez)
                self.entrez2cx[entrez] = cxid


        ws_vib2 = wb_vib.get_sheet_by_name(u'17971x 39')
        self.cx2row = {}
        self.cx2exp = {}
        for i in range(4, ws_vib2.max_row):
            cxid = ws_vib2['A'+str(i)].value
            self.cx2row[cxid] = i
            self.cx2exp[cxid] = {}
            self.cx2exp[cxid][0] = 0.0

            start_pos = 11
            exp_sum = 0.0
            for j in range(1, 40):
                exp = float(ws_vib2.cell(row=i,column=j+start_pos).value)
                self.cx2exp[cxid][j] = exp
                exp_sum += exp
                                
            # normalize
            '''
            for j in range(1, 40):
                exp = float(ws_vib2.cell(row=i,column=j+start_pos).value)
                self.cx2exp[cxid][j] = exp / exp_sum
            '''
        
        
        wb_tx = openpyxl.load_workbook(xlsfilename_tx)
        ws_tx = wb_tx.get_sheet_by_name(u'BAH用ISH画像リスト')
        self.cd2entrez = {}
        for i in range(2, ws_tx.max_row):
            if ws_tx['H'+str(i)].value is not None and ws_tx['H'+str(i)].value != u'':
                self.cd2entrez[ws_tx['A'+str(i)].value] = int(ws_tx['H'+str(i)].value)
        
        #print self.cx2entrez
        #print self.cd2entrez


    def convert_cd2cx(self):
        result = []
        for cdid, entrez in self.cd2entrez.items():
            record = {}
            if entrez in self.entrez2cx:
                cxid = self.entrez2cx[entrez]
            else:
                print 'No Entrez ID [%s]' % entrez
                continue
            
            if cxid in self.cx2row:
                cx_row =  self.cx2row[cxid]
                cx_exp = self.cx2exp[cxid]
            else:
                print 'No Cx ID [%s]' % cxid
                continue

            record['entrez'] = entrez
            record['cdid'] = cdid
            record['cxid'] = cxid
            record['cx_row'] = cx_row
            record['cx_exp'] = cx_exp

            #print record
            result.append(record)
            
        return result
        #return self.col[cd_id]
        
        
    def save_expression(self, result, filename):
        with open(filename, 'w') as f:
            for record in result:
                f.write('%s' % record['entrez'])
                for i in range(0, 40):
                    f.write(', %f' % record['cx_exp'][i])
                f.write('\n')

                
if __name__ == '__main__':
    cnv = ConvertID(xlsfilename_vib='../private/vibrism.xlsx', xlsfilename_tx='../private/braintx.xlsx')
    result = cnv.convert_cd2cx()
    #print result
    cnv.save_expression(result, '../private/cx_expression.txt')